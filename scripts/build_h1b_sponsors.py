"""Build the H-1B sponsor index from US DoL OFLC LCA disclosure files.

The DoL publishes quarterly LCA disclosure workbooks (free, official) at
https://www.dol.gov/agencies/eta/foreign-labor/performance — one big .xlsx per quarter, ~1M
rows, 75+ columns. This is a *build-time* ETL: download one or more of those workbooks, then::

    python scripts/build_h1b_sponsors.py ~/Downloads/LCA_Disclosure_Data_FY2025_Q*.xlsx

It streams the rows (openpyxl read-only, so the giant file never loads fully into memory), keeps
employers whose CASE_STATUS is *certified*, normalizes the employer name the same way dedup does
(so "STRIPE, INC." -> "stripe" matches a "Stripe" posting), counts certified filings, and writes
a compact ``src/ergon_tracker/registry/data/h1b_sponsors.json`` consumed by ``extract/visa.py``.

The parsing logic (``sponsors_from_rows``) is a pure function over row dicts, so it is unit-tested
without needing a real workbook.
"""

from __future__ import annotations

import json
import sys
from collections.abc import Iterable, Iterator
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from ergon_tracker.dedup import normalize_company  # noqa: E402

OUT = ROOT / "src" / "ergon_tracker" / "registry" / "data" / "h1b_sponsors.json"

# Header names vary slightly across fiscal years; match case-insensitively against these.
_EMPLOYER_COLS = ("EMPLOYER_NAME", "EMPLOYER_BUSINESS_DBA", "EMPLOYER_LEGAL_BUSINESS_NAME")
_STATUS_COLS = ("CASE_STATUS",)
# Filing-recency columns, in preference order (so users can spot sponsors that have gone quiet).
_DATE_COLS = ("DECISION_DATE", "RECEIVED_DATE", "CASE_SUBMITTED")
# Only certified LCAs count as demonstrated sponsorship ("Certified", "Certified - Withdrawn").
_CERTIFIED_PREFIX = "certified"


def _pick(row: dict[str, object], candidates: tuple[str, ...]) -> object | None:
    lowered = {k.lower(): v for k, v in row.items() if isinstance(k, str)}
    for cand in candidates:
        if cand.lower() in lowered:
            return lowered[cand.lower()]
    return None


def _to_iso(value: object) -> str | None:
    """Normalize an LCA date cell (datetime/date or string) to an ISO 'YYYY-MM-DD' string."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str) and value.strip():
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        return s[:10] if s[:4].isdigit() else None  # already ISO-ish
    return None


def sponsors_from_rows(rows: Iterable[dict[str, object]]) -> dict[str, dict[str, object]]:
    """Pure: rows -> {normalized_employer: {"n": certified_count, "last": latest_filing_iso}}."""
    out: dict[str, dict[str, object]] = {}
    for row in rows:
        status = _pick(row, _STATUS_COLS)
        if not isinstance(status, str) or not status.strip().lower().startswith(_CERTIFIED_PREFIX):
            continue
        employer = _pick(row, _EMPLOYER_COLS)
        if not isinstance(employer, str) or not employer.strip():
            continue
        key = normalize_company(employer)
        if not key:
            continue
        rec = out.setdefault(key, {"n": 0, "last": None})
        rec["n"] = int(rec["n"]) + 1  # type: ignore[arg-type]
        iso = _to_iso(_pick(row, _DATE_COLS))
        if iso and (rec["last"] is None or iso > str(rec["last"])):
            rec["last"] = iso  # keep the most recent filing date (ISO sorts lexicographically)
    return out


def read_xlsx_rows(path: Path) -> Iterator[dict[str, object]]:
    """Stream rows of an LCA workbook as header->value dicts (read-only, memory-bounded)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows)]
        for values in rows:
            yield dict(zip(header, values, strict=False))
    finally:
        wb.close()


def main(paths: list[str]) -> None:
    if not paths:
        print(__doc__)
        print("\nUsage: python scripts/build_h1b_sponsors.py <LCA_Disclosure_*.xlsx> [more ...]")
        raise SystemExit(2)

    total: dict[str, dict[str, object]] = {}
    for p in paths:
        path = Path(p).expanduser()
        if not path.exists():
            print(f"  ! missing: {path}")
            continue
        print(f"  reading {path.name} ...")
        part = sponsors_from_rows(read_xlsx_rows(path))
        for name, rec in part.items():
            agg = total.setdefault(name, {"n": 0, "last": None})
            agg["n"] = int(agg["n"]) + int(rec["n"])  # type: ignore[arg-type]
            last = rec["last"]
            if last and (agg["last"] is None or str(last) > str(agg["last"])):
                agg["last"] = last
        print(f"    +{len(part):,} employers (running total {len(total):,})")

    payload = {
        "source": "US DoL OFLC LCA disclosure data (certified filings)",
        "count": len(total),
        # name -> {"n": certified filing count, "last": most-recent filing date (ISO)}.
        "sponsors": dict(sorted(total.items())),
    }
    OUT.write_text(json.dumps(payload), encoding="utf-8")
    print(f"\nwrote {OUT.relative_to(ROOT)} — {len(total):,} unique sponsors")


if __name__ == "__main__":
    main(sys.argv[1:])
